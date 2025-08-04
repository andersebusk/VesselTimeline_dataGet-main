import msal
import requests
import openpyxl
import re
import unicodedata
import warnings
from datetime import datetime as dt
import difflib
from io import BytesIO
import pandas as pd
import numpy as np
import os

# ==============================
# CONFIGURATION
# ==============================
TENANT_ID = os.getenv("TENANT_ID")
CLIENT_ID = os.getenv("CLIENT_ID")
CLIENT_SECRET = os.getenv("CLIENT_SECRET")
TARGET_SITE_DISPLAY_NAME = os.getenv("TARGET_SITE_DISPLAY_NAME")
FOLDER_PATH = os.getenv("FOLDER_PATH")
TARGET_FILE_NAME = os.getenv("TARGET_FILE_NAME")

PBI_WORKSPACE_ID = os.getenv("PBI_WORKSPACE_ID")
PBI_TENANT_ID = os.getenv("PBI_TENANT_ID")
PBI_CLIENT_ID = os.getenv("PBI_CLIENT_ID")
PBI_CLIENT_SECRET = os.getenv("PBI_CLIENT_SECRET")
TARGET_DATASET_NAME = "VesselData"  # Power BI streaming dataset name

# ==============================
# HELPER FUNCTIONS
# ==============================
def normalize_string(s):
    s = s.lower().strip()
    s = unicodedata.normalize('NFKD', s).encode('ascii', 'ignore').decode('utf-8', 'ignore')
    s = re.sub(r'[^\w\s]', ' ', s)
    s = re.sub(r'\s+', ' ', s).strip()
    return s

def string_similarity(str1, str2):
    return difflib.SequenceMatcher(None, normalize_string(str1), normalize_string(str2)).ratio()

def find_value_columns_by_headers(sheet, header_strings):
    header_indices = {}
    for row in sheet.iter_rows(min_row=4, max_row=4):
        for index, cell in enumerate(row[:100]):
            cell_value_str = str(cell.value).strip() if cell.value else ""
            for target_header in header_strings:
                target_normalized = normalize_string(target_header)
                if target_header.startswith(("Fe magnetic", "Fe corrosive", "Fe total", "Residual TBN", "Unit")):
                    if target_normalized == normalize_string(cell_value_str):
                        header_indices[target_header] = index
                        break
                elif string_similarity(cell_value_str, target_normalized) >= 0.60:
                    header_indices[target_header] = index
    return header_indices

def parse_date(date_value):
    if isinstance(date_value, dt):
        return date_value.strftime("%Y-%m-%d")
    date_patterns = [r"(\d{1,2})[\/\-.](\d{1,2})[\/\-.](\d{2,4})"]
    for pattern in date_patterns:
        match = re.match(pattern, str(date_value))
        if match:
            day, month, year = match.groups()
            if len(year) == 2:
                year = '20' + year
            try:
                date_obj = dt.strptime(f"{day}-{month}-{year}", "%d-%m-%Y")
                return date_obj.strftime("%Y-%m-%d")
            except ValueError:
                return None
    return None

def map_sheet_names(workbook):
    mapping = {}
    overview_sheet = workbook["Overview"]
    for row in overview_sheet.iter_rows(min_row=2, max_row=150, values_only=True):
        if row[4] is not None and row[3] is not None:
            try:
                mapping[int(row[4])] = row[3]
            except ValueError:
                mapping[str(row[4])] = row[3]
    return mapping

def detect_max_cylinders(header_columns):
    max_unit = 0
    for key in header_columns.keys():
        match = re.match(r".*?(\d+)$", key)
        if match:
            max_unit = max(max_unit, int(match.group(1)))
    return max_unit

def process_xlsx(sheet, date_col, header_columns, max_rows=None):
    data_rows = []
    max_cylinders = detect_max_cylinders(header_columns)

    for row in sheet.iter_rows(min_row=6, max_row=max_rows or sheet.max_row, values_only=True):
        # --- Parse Date ---
        date_value = row[date_col]
        if not date_value:
            continue
        parsed_date = parse_date(date_value)
        if not parsed_date:
            continue

        # --- Extract ME Load, ME RH ---
        me_load = row[1] if len(row) > 1 else None
        me_rh_col = header_columns.get("ME rh", header_columns.get("ME", None))
        me_rh = row[me_rh_col] if me_rh_col is not None and me_rh_col < len(row) else None

        # --- Extract TBN Fed & Fuel Sulphur ---
        tbn_fed_index = header_columns.get("TBN of blended oil fed to engine", None)
        tbn_fed_value = row[tbn_fed_index] if tbn_fed_index is not None and tbn_fed_index < len(row) else None

        fuel_sulphur_index = header_columns.get("Fuel Sulphur Content", None)
        fuel_sulphur_value = row[fuel_sulphur_index] if fuel_sulphur_index is not None and fuel_sulphur_index < len(row) else None

        # --- Loop through cylinders ---
        for unit_num in range(1, max_cylinders + 1):
            fe_magnetic_col = header_columns.get(f"Fe magnetic {unit_num}")
            fe_corrosive_col = header_columns.get(f"Fe corrosive {unit_num}")
            fe_total_col = header_columns.get(f"Fe total {unit_num}")
            residual_tbn_col = header_columns.get(f"Residual TBN {unit_num}")
            unit_col = header_columns.get(f"Unit {unit_num}")

            fe_magnetic_val = row[fe_magnetic_col] if fe_magnetic_col is not None and fe_magnetic_col < len(row) else None
            fe_corrosive_val = row[fe_corrosive_col] if fe_corrosive_col is not None and fe_corrosive_col < len(row) else None
            fe_total_val = row[fe_total_col] if fe_total_col is not None and fe_total_col < len(row) else None
            residual_tbn_val = row[residual_tbn_col] if residual_tbn_col is not None and residual_tbn_col < len(row) else None
            unit_val = row[unit_col] if unit_col is not None and unit_col < len(row) else None

            data_rows.append({
                "Date": parsed_date,
                "Cylinder": f"Cyl. {unit_num}",
                "ME_RH": me_rh,
                "Fe_Magnet": fe_magnetic_val,
                "Fe_Corrosion": fe_corrosive_val,
                "Fe_Total": fe_total_val,
                "Residual_TBN": residual_tbn_val,
                "Unit_RH": unit_val,
                "TBN_Fed": None,
                "Fuel_Sulph": fuel_sulphur_value,
                "ME_Load": None
            })

        # --- Add TBN Fed Row ---
        data_rows.append({
            "Date": parsed_date,
            "Cylinder": "TBN Fed",
            "ME_RH": me_rh,
            "Fe_Magnet": None,
            "Fe_Corrosion": None,
            "Fe_Total": None,
            "Residual_TBN": None,
            "Unit_RH": None,
            "TBN_Fed": tbn_fed_value,
            "Fuel_Sulph": fuel_sulphur_value,
            "ME_Load": None
        })

        # --- Add ME Load Row ---
        data_rows.append({
            "Date": parsed_date,
            "Cylinder": "ME Load",
            "ME_RH": me_rh,
            "Fe_Magnet": None,
            "Fe_Corrosion": None,
            "Fe_Total": None,
            "Residual_TBN": None,
            "Unit_RH": None,
            "TBN_Fed": None,
            "Fuel_Sulph": fuel_sulphur_value,
            "ME_Load": me_load
        })

    return data_rows


# ==============================
# STEP 1: AUTHENTICATE & DOWNLOAD EXCEL
# ==============================
print("🔑 Authenticating to Microsoft Graph...")
app = msal.ConfidentialClientApplication(CLIENT_ID, authority=f"https://login.microsoftonline.com/{TENANT_ID}", client_credential=CLIENT_SECRET)
token = app.acquire_token_for_client(scopes=["https://graph.microsoft.com/.default"])
headers = {"Authorization": f"Bearer {token['access_token']}"}

site_resp = requests.get(f"https://graph.microsoft.com/v1.0/sites?search={TARGET_SITE_DISPLAY_NAME}", headers=headers)
site_id = site_resp.json()["value"][0]["id"]

files_resp = requests.get(f"https://graph.microsoft.com/v1.0/sites/{site_id}/drive/root:/{FOLDER_PATH}:/children", headers=headers).json()
file_id = next(f["id"] for f in files_resp["value"] if f["name"].lower() == TARGET_FILE_NAME.lower())

file_dl_url = f"https://graph.microsoft.com/v1.0/sites/{site_id}/drive/items/{file_id}/content"
file_resp = requests.get(file_dl_url, headers=headers)
workbook = openpyxl.load_workbook(BytesIO(file_resp.content), data_only=True, read_only=True)
print("✅ Excel downloaded successfully!")

# ==============================
# STEP 2: EXTRACT AND FORMAT DATA
# ==============================
headers_to_find = (
    [f"Fe magnetic {i}" for i in range(1, 13)] +
    [f"Fe corrosive {i}" for i in range(1, 13)] +
    [f"Fe total {i}" for i in range(1, 13)] +
    [f"Residual TBN {i}" for i in range(1, 13)] +
    [f"Unit {i}" for i in range(1, 13)] +
    ["ME rh", "ME", "TBN of blended oil fed to engine", "Fuel Sulphur Content"]
)

sheet_mapping = map_sheet_names(workbook)
all_rows = []

for sheet_name in [s for s in workbook.sheetnames if s not in ["Overview", "Dashboard table", "Dashboard"]]:
    sheet = workbook[sheet_name]
    header_columns = find_value_columns_by_headers(sheet, headers_to_find)
    if not header_columns:
        print(f"⚠️ Skipping '{sheet_name}' (headers not found)")
        continue

    rows = process_xlsx(sheet, 0, header_columns)

    try:
        sheet_key = int(sheet_name)
    except ValueError:
        sheet_key = sheet_name

    vessel_name = sheet_mapping.get(sheet_key)
    if not vessel_name:
        print(f"⚠️ Skipping '{sheet_name}' (no vessel mapping found)")
        continue

    for r in rows:
        r["VesselID"] = vessel_name
        all_rows.append(r)

df = pd.DataFrame(all_rows)
print(f"✅ Extracted {len(df)} rows for Power BI push.")

# Clean numeric fields
numeric_columns = ["Fe_Magnet", "Fe_Corrosion", "Fe_Total", "Residual_TBN", "Unit_RH", "TBN_Fed", "Fuel_Sulph", "ME_Load", "ME_RH"]
for col in numeric_columns:
    df[col] = pd.to_numeric(df[col], errors="coerce")

df["Date"] = pd.to_datetime(df["Date"], errors="coerce").dt.strftime('%Y-%m-%d')

# ==============================
# APPLY RENAMING FOR POWER BI
# ==============================
print("🔄 Renaming columns for Power BI output...")

PBI_COLUMN_MAPPING = {
    "VesselID": "Vessel Name",
    "Date": "Date",
    "Cylinder": "Cylinder",
    "Fe_Magnet": "Fe Magnetic",
    "Fe_Corrosion": "Fe Corrosive",
    "Fe_Total": "Fe Total",
    "Residual_TBN": "Residual TBN",
    "Unit_RH": "Unit RH",
    "TBN_Fed": "TBN Fed",
    "Fuel_Sulph": "Fuel Sulphur Content",
    "ME_Load": "ME Load",
    "ME_RH": "ME RH"
}

df.rename(columns=PBI_COLUMN_MAPPING, inplace=True)

# ==============================
# STEP 3: PUSH TO POWER BI
# ==============================
print("🔑 Authenticating Power BI...")
pbi_app = msal.ConfidentialClientApplication(PBI_CLIENT_ID, authority=f"https://login.microsoftonline.com/{PBI_TENANT_ID}", client_credential=PBI_CLIENT_SECRET)
pbi_token = pbi_app.acquire_token_for_client(scopes=["https://analysis.windows.net/powerbi/api/.default"])
pbi_headers = {"Authorization": f"Bearer {pbi_token['access_token']}"}

datasets_url = f"https://api.powerbi.com/v1.0/myorg/groups/{PBI_WORKSPACE_ID}/datasets"
datasets = requests.get(datasets_url, headers=pbi_headers).json().get("value", [])
dataset = next((d for d in datasets if d.get("addRowsAPIEnabled") and d["name"].lower() == TARGET_DATASET_NAME.lower()), None)
if not dataset:
    raise SystemExit(f"❌ Streaming dataset '{TARGET_DATASET_NAME}' not found.")

PBI_DATASET_ID = dataset["id"]
tables_url = f"https://api.powerbi.com/v1.0/myorg/groups/{PBI_WORKSPACE_ID}/datasets/{PBI_DATASET_ID}/tables"
PBI_TABLE_NAME = requests.get(tables_url, headers=pbi_headers).json()["value"][0]["name"]

# Clear old rows
pbi_clear_url = f"https://api.powerbi.com/v1.0/myorg/groups/{PBI_WORKSPACE_ID}/datasets/{PBI_DATASET_ID}/tables/{PBI_TABLE_NAME}/rows"
print("🗑 Clearing old rows in Power BI...")
requests.delete(pbi_clear_url, headers=pbi_headers)

# Push rows
rows_to_push = df.replace([np.nan, np.inf, -np.inf], None).to_dict(orient="records")
print(f"📤 Pushing {len(rows_to_push)} rows to Power BI...")
for i in range(0, len(rows_to_push), 10000):
    batch = rows_to_push[i:i+10000]
    resp = requests.post(pbi_clear_url, headers={**pbi_headers, "Content-Type": "application/json"}, json={"rows": batch})
    if resp.status_code not in [200, 202]:
        raise SystemExit(f"❌ Failed to push batch {i//10000+1}: {resp.status_code} {resp.text}")

print("🎉 All data successfully pushed to Power BI Streaming Dataset!")
