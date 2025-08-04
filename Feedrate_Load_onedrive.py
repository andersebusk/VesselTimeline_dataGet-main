import msal
import requests
import openpyxl
import re
import unicodedata
import warnings
from datetime import datetime as dt
import difflib
from io import BytesIO
import pandas as pd
import os
import numpy as np

# ==============================
# CONFIGURATION
# ==============================
TENANT_ID = os.getenv("TENANT_ID")
CLIENT_ID = os.getenv("CLIENT_ID")
CLIENT_SECRET = os.getenv("CLIENT_SECRET")
TARGET_SITE_DISPLAY_NAME = os.getenv("TARGET_SITE_DISPLAY_NAME")
FOLDER_PATH = os.getenv("FOLDER_PATH")  # Folder in OneDrive/SharePoint
TARGET_FILE_NAME = os.getenv("TARGET_FILE_NAME")  # Excel filename

PBI_WORKSPACE_ID = os.getenv("PBI_WORKSPACE_ID")
PBI_TENANT_ID = os.getenv("PBI_TENANT_ID")
PBI_CLIENT_ID = os.getenv("PBI_CLIENT_ID")
PBI_CLIENT_SECRET = os.getenv("PBI_CLIENT_SECRET")
TARGET_DATASET_NAME = "FeedrateData"  # Power BI streaming dataset name

# ==============================
# HELPER FUNCTIONS
# ==============================
def normalize_string(s):
    s = s.lower().strip()
    s = unicodedata.normalize('NFKD', s).encode('ascii', 'ignore').decode('utf-8', 'ignore')
    s = re.sub(r'[^\w\s]', ' ', s)
    s = re.sub(r'\s+', ' ', s)
    return s.strip()

def string_similarity(str1, str2):
    return difflib.SequenceMatcher(None, normalize_string(str1), normalize_string(str2)).ratio()

def find_value_columns_by_headers(sheet, header_strings):
    header_indices = {}
    try:
        for row in sheet.iter_rows(min_row=4, max_row=4):
            for index, cell in enumerate(row[:100]):
                cell_value_str = str(cell.value) if cell.value else ""
                cell_value_str_normalized = normalize_string(cell_value_str)
                for target_header in header_strings:
                    target_normalized = normalize_string(target_header)
                    if string_similarity(cell_value_str_normalized, target_normalized) >= 0.60:
                        header_indices[target_header] = index
        return {k: v for k, v in header_indices.items() if v is not None}
    except Exception as e:
        print(f"❌ Error finding headers: {e}")
    return {}

def parse_date(date_value):
    if isinstance(date_value, dt):
        return date_value.strftime("%Y-%m-%d")
    date_patterns = [r"(\d{1,2})[\/\-.](\d{1,2})[\/\-.](\d{2,4})"]
    for pattern in date_patterns:
        match = re.match(pattern, str(date_value))
        if match:
            part1, part2, year = match.groups()
            if len(year) == 2:
                year = '20' + year
            try:
                date_obj = dt.strptime(f"{part1}-{part2}-{year}", "%d-%m-%Y")
                return date_obj.strftime("%Y-%m-%d")
            except ValueError:
                return None
    return None

def map_sheet_names(workbook):
    mapping = {}
    overview_sheet = workbook["Overview"]
    for row in overview_sheet.iter_rows(min_row=2, max_row=150, values_only=True):
        if row[4] is not None and row[3] is not None:
            try:
                mapping[int(row[4])] = row[3]
            except ValueError:
                mapping[str(row[4])] = row[3]
    return mapping

def process_xlsx(sheet, date_column_index, ME_load_index, header_columns, vessel_name, max_rows_to_search=None):
    rows = []
    max_row = max_rows_to_search if max_rows_to_search else sheet.max_row
    for row in sheet.iter_rows(min_row=6, max_row=max_row, values_only=True):
        date_value = row[date_column_index]
        if not date_value:
            continue
        parsed_date = parse_date(date_value)
        if not parsed_date:
            continue
        
        me_rh_value = None
        if header_columns.get("ME rh") is not None:
            me_rh_value = row[header_columns["ME rh"]]
        elif header_columns.get("ME") is not None:
            me_rh_value = row[header_columns["ME"]]

        row_data = {
            "VesselID": vessel_name,
            "Date": parsed_date,
            "ME_Load": row[ME_load_index] or 0,
            "CylinderOilFeedrate": row[header_columns.get("Cylinder oil feedrate")] if header_columns.get("Cylinder oil feedrate") is not None else 0,
            "ME_RH": me_rh_value or 0
        }
        rows.append(row_data)
    return rows

# ==============================
# STEP 1: AUTHENTICATE & DOWNLOAD FILE FROM ONEDRIVE
# ==============================
print("🔑 Authenticating to Microsoft Graph...")
app = msal.ConfidentialClientApplication(
    CLIENT_ID, 
    authority=f"https://login.microsoftonline.com/{TENANT_ID}", 
    client_credential=CLIENT_SECRET
)
token = app.acquire_token_for_client(scopes=["https://graph.microsoft.com/.default"])
access_token = token["access_token"]
headers = {"Authorization": f"Bearer {access_token}"}

# Fetch site and file
site_resp = requests.get(f"https://graph.microsoft.com/v1.0/sites?search={TARGET_SITE_DISPLAY_NAME}", headers=headers)
site_id = site_resp.json()["value"][0]["id"]
folder_url = f"https://graph.microsoft.com/v1.0/sites/{site_id}/drive/root:/{FOLDER_PATH}:/children"
files_resp = requests.get(folder_url, headers=headers).json()
file_id = next(f["id"] for f in files_resp["value"] if f["name"].lower() == TARGET_FILE_NAME.lower())
file_dl_url = f"https://graph.microsoft.com/v1.0/sites/{site_id}/drive/items/{file_id}/content"

print("📥 Downloading Excel from OneDrive...")
file_resp = requests.get(file_dl_url, headers=headers)
workbook = openpyxl.load_workbook(BytesIO(file_resp.content), data_only=True, read_only=True)
print("✅ Excel downloaded successfully!")

# ==============================
# STEP 2: EXTRACT & FORMAT DATA
# ==============================
headers_to_find = ['ME load', 'Cylinder oil feedrate', 'ME rh', 'ME']
sheet_mapping = map_sheet_names(workbook)
all_rows = []

for sheet_name in [s for s in workbook.sheetnames if s not in ["Overview", "Dashboard table", "Dashboard"]]:
    sheet = workbook[sheet_name]
    header_columns = find_value_columns_by_headers(sheet, headers_to_find)
    if not header_columns:
        print(f"⚠️ Skipping '{sheet_name}' (headers not found)")
        continue

    try:
        sheet_key = int(sheet_name)
    except ValueError:
        sheet_key = sheet_name
    vessel_name = sheet_mapping.get(sheet_key, None)
    if not vessel_name:
        print(f"⚠️ Skipping '{sheet_name}' (vessel mapping not found)")
        continue

    rows = process_xlsx(sheet, 0, 1, header_columns, vessel_name, max_rows_to_search=170)
    all_rows.extend(rows)
    print(f"✅ Processed '{sheet_name}' → {len(rows)} rows added")

print(f"\n📊 Total rows ready for Power BI push: {len(all_rows)}")

# Convert to DataFrame
df = pd.DataFrame(all_rows)

# Clean numeric fields
numeric_columns = ["ME_Load", "CylinderOilFeedrate", "ME_RH"]
for col in numeric_columns:
    df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)

# Ensure Date is ISO for Power BI
df["Date"] = pd.to_datetime(df["Date"], errors="coerce").dt.strftime('%Y-%m-%d')
print("✅ Data cleaned and ready for Power BI push.")

# ==============================
# APPLY RENAMING FOR POWER BI
# ==============================
print("🔄 Renaming columns for Power BI output...")

PBI_COLUMN_MAPPING = {
   "VesselID": "Vessel Name",
    "Date": "Date",
    "ME_Load": "ME Load",
    "CylinderOilFeedrate": "Cylinder Oil Feedrate",
    "ME_RH": "ME RH"
}

df.rename(columns=PBI_COLUMN_MAPPING, inplace=True)

# ==============================
# STEP 3: AUTHENTICATE POWER BI AND PUSH DATA
# ==============================
print("🔑 Authenticating Power BI Service Principal...")
pbi_app = msal.ConfidentialClientApplication(
    client_id=PBI_CLIENT_ID,
    authority=f"https://login.microsoftonline.com/{PBI_TENANT_ID}",
    client_credential=PBI_CLIENT_SECRET,
)
pbi_token = pbi_app.acquire_token_for_client(scopes=["https://analysis.windows.net/powerbi/api/.default"])

if "access_token" not in pbi_token:
    print("❌ Failed to acquire Power BI token:", pbi_token.get("error_description", pbi_token))
    raise SystemExit("Stopping: Authentication failed.")

pbi_access_token = pbi_token["access_token"]
pbi_headers = {"Authorization": f"Bearer {pbi_access_token}"}

# Verify dataset and table
datasets_url = f"https://api.powerbi.com/v1.0/myorg/groups/{PBI_WORKSPACE_ID}/datasets"
datasets = requests.get(datasets_url, headers=pbi_headers).json().get("value", [])
dataset = next((ds for ds in datasets if ds.get("addRowsAPIEnabled") and ds["name"].lower() == TARGET_DATASET_NAME.lower()), None)
if not dataset:
    raise SystemExit(f"❌ Streaming dataset '{TARGET_DATASET_NAME}' not found in workspace.")

PBI_DATASET_ID = dataset["id"]
tables_url = f"https://api.powerbi.com/v1.0/myorg/groups/{PBI_WORKSPACE_ID}/datasets/{PBI_DATASET_ID}/tables"
PBI_TABLE_NAME = requests.get(tables_url, headers=pbi_headers).json()["value"][0]["name"]

# Convert dataframe to JSON-friendly records
rows_to_push = df.replace([np.nan, np.inf, -np.inf], None).to_dict(orient="records")

# Clear old rows in Power BI dataset
pbi_clear_url = f"https://api.powerbi.com/v1.0/myorg/groups/{PBI_WORKSPACE_ID}/datasets/{PBI_DATASET_ID}/tables/{PBI_TABLE_NAME}/rows"
print("🗑 Clearing old rows in Power BI streaming dataset...")
requests.delete(pbi_clear_url, headers=pbi_headers)

# Push rows to Power BI
print(f"📤 Pushing {len(rows_to_push)} rows to Power BI...")
batch_size = 10000
for i in range(0, len(rows_to_push), batch_size):
    batch = rows_to_push[i:i + batch_size]
    push_response = requests.post(pbi_clear_url, headers={**pbi_headers, "Content-Type": "application/json"}, json={"rows": batch})
    if push_response.status_code in [200, 202]:
        print(f"✅ Batch {i//batch_size+1} pushed successfully ({len(batch)} rows).")
    else:
        print(f"❌ Failed to push batch {i//batch_size+1}: {push_response.status_code} {push_response.text}")
        raise SystemExit("Stopping due to API error.")

print("🎉 All data successfully pushed to Power BI Streaming Dataset!")
