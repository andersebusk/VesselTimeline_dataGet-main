import msal
import requests
import openpyxl
import re
import unicodedata
import warnings
from datetime import datetime as dt
import difflib
from io import BytesIO
import pandas as pd
import os
import numpy as np

# ==============================
# CONFIGURATION
# ==============================
TENANT_ID = os.getenv("TENANT_ID")
CLIENT_ID = os.getenv("CLIENT_ID")
CLIENT_SECRET = os.getenv("CLIENT_SECRET")
TARGET_SITE_DISPLAY_NAME = os.getenv("TARGET_SITE_DISPLAY_NAME")
FOLDER_PATH = os.getenv("FOLDER_PATH_ME")
TARGET_FILE_NAME = os.getenv("TARGET_FILE_NAME_ME")

PBI_WORKSPACE_ID = os.getenv("PBI_WORKSPACE_ID")
PBI_TENANT_ID = os.getenv("PBI_TENANT_ID")
PBI_CLIENT_ID = os.getenv("PBI_CLIENT_ID")
PBI_CLIENT_SECRET = os.getenv("PBI_CLIENT_SECRET")
TARGET_DATASET_NAME = "MeSYSoil"

HEADERS_TO_FIND = [
    "Viscosity @ 40C",
    "Viscosity @ 100C",
    "Base Number",
    "Water Level",
    "Top up volume",
    "PQ Index",
    "Oil on label",
    "Iso Code",
    "Particle count > 4 [μm/ml]",
    "Particle count > 6 [μm/ml]",
    "Particle count > 14 [μm/ml]",
    "Vanadium"
]

# ==============================
# HELPER FUNCTIONS
# ==============================
def normalize_string(s):
    s = s.lower().strip()
    s = unicodedata.normalize('NFKD', s).encode('ascii', 'ignore').decode('utf-8', 'ignore')
    s = re.sub(r'[^\w\s]', ' ', s)
    s = re.sub(r'\s+', ' ', s).strip()
    return s

def string_similarity(str1, str2):
    return difflib.SequenceMatcher(None, normalize_string(str1), normalize_string(str2)).ratio()

def find_headers(sheet, header_strings):
    header_indices = {}
    for row in sheet.iter_rows(min_row=4, max_row=4):
        for index, cell in enumerate(row[:100]):
            cell_value_str = str(cell.value).strip() if cell.value else ""
            for target_header in header_strings:
                if string_similarity(cell_value_str, target_header) >= 0.60:
                    header_indices[target_header] = index
    return header_indices

def parse_date(date_value):
    if isinstance(date_value, dt):
        return date_value.strftime("%Y-%m-%d")
    date_patterns = [r"(\d{1,2})[\/\-.](\d{1,2})[\/\-.](\d{2,4})"]
    for pattern in date_patterns:
        match = re.match(pattern, str(date_value))
        if match:
            day, month, year = match.groups()
            if len(year) == 2:
                year = '20' + year
            try:
                date_obj = dt.strptime(f"{day}-{month}-{year}", "%d-%m-%Y")
                return date_obj.strftime("%Y-%m-%d")
            except ValueError:
                return None
    return None

def extract_me_sys_data(workbook, headers_to_find):
    data_rows = []
    warnings.simplefilter("ignore")
    sheet_names = [s for s in workbook.sheetnames if s not in ["Overview", "Dashboard table", "Dashboard"]]

    for sheet_name in sheet_names:
        sheet = workbook[sheet_name]
        date_col = 0
        header_columns = find_headers(sheet, headers_to_find)
        if not header_columns:
            print(f"⚠️ Skipping sheet '{sheet_name}': No headers found.")
            continue

        for row in sheet.iter_rows(min_row=6, max_row=sheet.max_row, values_only=True):
            date_value = row[date_col]
            if not date_value:
                continue
            parsed_date = parse_date(date_value)
            if not parsed_date:
                continue

            row_data = {"VesselID": sheet_name, "Date": parsed_date}
            for header in headers_to_find:
                col_index = header_columns.get(header)
                value = row[col_index] if col_index is not None and col_index < len(row) else None
                row_data[header] = value
            data_rows.append(row_data)
        print(f"✅ Processed sheet: {sheet_name} ({len(data_rows)} total rows so far)")
    return pd.DataFrame(data_rows)

# ==============================
# STEP 1: FETCH EXCEL FROM ONEDRIVE
# ==============================
print("🔑 Authenticating to Microsoft Graph...")
app = msal.ConfidentialClientApplication(CLIENT_ID, authority=f"https://login.microsoftonline.com/{TENANT_ID}", client_credential=CLIENT_SECRET)
token = app.acquire_token_for_client(scopes=["https://graph.microsoft.com/.default"])
access_token = token["access_token"]
headers = {"Authorization": f"Bearer {access_token}"}

site_resp = requests.get(f"https://graph.microsoft.com/v1.0/sites?search={TARGET_SITE_DISPLAY_NAME}", headers=headers)
site_id = site_resp.json()["value"][0]["id"]

files_resp = requests.get(f"https://graph.microsoft.com/v1.0/sites/{site_id}/drive/root:/{FOLDER_PATH}:/children", headers=headers).json()
file_id = next(f["id"] for f in files_resp["value"] if f["name"].lower() == TARGET_FILE_NAME.lower())

print("📥 Downloading Excel from OneDrive...")
file_dl_url = f"https://graph.microsoft.com/v1.0/sites/{site_id}/drive/items/{file_id}/content"
file_resp = requests.get(file_dl_url, headers=headers)
workbook = openpyxl.load_workbook(BytesIO(file_resp.content), data_only=True, read_only=True)
print("✅ Excel downloaded successfully!")

# ==============================
# STEP 2: EXTRACT DATA
# ==============================
print("📊 Extracting ME SYS Oil data...")
df = extract_me_sys_data(workbook, HEADERS_TO_FIND)

# Map headers to clean names
header_mapping = {
    "Viscosity @ 100C": "KVisc100",
    "Base Number": "BN",
    "Top up volume": "TopUPVolume",
    "PQ Index": "PQIndex",
    "Oil on label": "OilOnLabel",
    "Iso Code": "ISOCode",
    "Particle count > 4 [μm/ml]": "PartCount4",
    "Particle count > 6 [μm/ml]": "PartCount6",
    "Particle count > 14 [μm/ml]": "PartCount14",
    "Vanadium": "Vanadium"
}
df = df.rename(columns=header_mapping)

# Ensure required columns
db_columns = ["VesselID", "Date", "KVisc100", "BN", "TopUPVolume", "Vanadium", 
              "PQIndex", "OilOnLabel", "ISOCode", "PartCount4", "PartCount6", "PartCount14"]
for col in db_columns:
    if col not in df.columns:
        df[col] = None
df = df[db_columns]

# Clean numeric fields and format date
df["Date"] = pd.to_datetime(df["Date"], errors="coerce").dt.strftime('%Y-%m-%d')
for col in db_columns:
    if col not in ["VesselID", "Date"]:
        df[col] = pd.to_numeric(df[col], errors="coerce")

print(f"✅ Extracted {len(df)} rows ready for Power BI")

# ==============================
# APPLY RENAMING FOR POWER BI
# ==============================
print("🔄 Renaming columns for Power BI output...")

PBI_COLUMN_MAPPING = {
    "VesselID": "Vessel Name",
    "Date": "Date",
    "KVisc100": "Viscosity @ 100C",
    "BN": "BN",
    "TopUPVolume": "Top Up Volume",
    "Vanadium": "Vanadium",
    "PQIndex": "PQ Index",
    "OilOnLabel": "Oil On Label",
    "ISOCode": "ISO Code",
    "PartCount4": "Particle count > 4",
    "PartCount6": "Particle count > 6",
    "PartCount14": "Particle count > 14"
}

df.rename(columns=PBI_COLUMN_MAPPING, inplace=True)

# ✅ Validate mapped columns
missing_cols = [col for col in PBI_COLUMN_MAPPING.values() if col not in df.columns]
if missing_cols:
    print(f"⚠️ Warning: Missing Power BI columns in DataFrame: {missing_cols}")
else:
    print("✅ All Power BI columns renamed successfully.")

# ==============================
# STEP 3: PUSH TO POWER BI DIRECTLY
# ==============================
print("🔑 Authenticating Power BI Service Principal...")
pbi_app = msal.ConfidentialClientApplication(PBI_CLIENT_ID, authority=f"https://login.microsoftonline.com/{PBI_TENANT_ID}", client_credential=PBI_CLIENT_SECRET)
pbi_token = pbi_app.acquire_token_for_client(scopes=["https://analysis.windows.net/powerbi/api/.default"])
pbi_access_token = pbi_token["access_token"]
pbi_headers = {"Authorization": f"Bearer {pbi_access_token}"}

datasets_url = f"https://api.powerbi.com/v1.0/myorg/groups/{PBI_WORKSPACE_ID}/datasets"
datasets = requests.get(datasets_url, headers=pbi_headers).json().get("value", [])
dataset = next((ds for ds in datasets if ds.get("addRowsAPIEnabled") and ds["name"].lower() == TARGET_DATASET_NAME.lower()), None)
if not dataset:
    raise SystemExit(f"❌ Streaming dataset '{TARGET_DATASET_NAME}' not found in workspace.")

PBI_DATASET_ID = dataset["id"]
tables_url = f"https://api.powerbi.com/v1.0/myorg/groups/{PBI_WORKSPACE_ID}/datasets/{PBI_DATASET_ID}/tables"
PBI_TABLE_NAME = requests.get(tables_url, headers=pbi_headers).json()["value"][0]["name"]

# Prepare rows
rows_to_push = df.replace([np.nan, np.inf, -np.inf], None).to_dict(orient="records")

print("🛠 Preview 'Oil On Label' values being pushed to Power BI:")
for r in rows_to_push[:10]:  # Show first 10 rows for inspection
    print(f"Vessel: {r.get('Vessel Name')}, Date: {r.get('Date')}, Oil On Label: {r.get('Oil On Label')}")
# Clear old rows in streaming dataset
pbi_clear_url = f"https://api.powerbi.com/v1.0/myorg/groups/{PBI_WORKSPACE_ID}/datasets/{PBI_DATASET_ID}/tables/{PBI_TABLE_NAME}/rows"
print("🗑 Clearing old rows in Power BI...")
requests.delete(pbi_clear_url, headers=pbi_headers)

# Push rows
print(f"📤 Pushing {len(rows_to_push)} rows to Power BI...")
for i in range(0, len(rows_to_push), 10000):
    batch = rows_to_push[i:i+10000]
    resp = requests.post(pbi_clear_url, headers={**pbi_headers, "Content-Type": "application/json"}, json={"rows": batch})
    if resp.status_code not in [200, 202]:
        print(f"❌ Failed to push batch {i//10000+1}: {resp.status_code} {resp.text}")
        raise SystemExit("Stopping due to API error.")
print("🎉 Data successfully pushed to Power BI (no MySQL used)!")
